import sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Connect to database
conn = sqlite3.connect('backend/project/db.sqlite3')

# Query 1: Get all BU Objectives with org unit names
bu_objectives_query = """
SELECT 
    bo.id,
    ou.name as bu_name,
    bo.parameter_name,
    bo.goal_text,
    bo.measure_of_success,
    bo.linkage_ta_raw,
    bo.linkage_go_raw,
    bo.source_sheet,
    bo.source_row_no
FROM smart_hr_backend_buobjective bo
LEFT JOIN smart_hr_backend_orgunit ou ON bo.org_unit_id = ou.id
ORDER BY ou.name, bo.id
"""

# Query 2: Get all Thrust Areas
thrust_areas_query = """
SELECT 
    id,
    code,
    description,
    is_sub_heading,
    parent_code
FROM smart_hr_backend_thrustarea
ORDER BY code
"""

# Query 3: Get all Group Objectives
group_objectives_query = """
SELECT 
    id,
    code,
    description,
    parameter,
    is_sub_heading,
    parent_code
FROM smart_hr_backend_groupobjective
ORDER BY code
"""

# Query 4: Get BU Objective to Thrust Area linkages
bu_ta_links_query = """
SELECT 
    bta.id,
    ou.name as bu_name,
    bo.parameter_name,
    bta.ta_code_raw,
    bta.ta_code_normalized,
    ta.description as ta_description
FROM smart_hr_backend_buobjectivetalink bta
LEFT JOIN smart_hr_backend_buobjective bo ON bta.objective_id = bo.id
LEFT JOIN smart_hr_backend_orgunit ou ON bo.org_unit_id = ou.id
LEFT JOIN smart_hr_backend_thrustarea ta ON bta.ta_code_normalized = ta.code
ORDER BY ou.name, bo.id
"""

# Query 5: Get BU Objective to Group Objective linkages
bu_go_links_query = """
SELECT 
    bgo.id,
    ou.name as bu_name,
    bo.parameter_name,
    bgo.go_code_raw,
    bgo.go_code_normalized,
    go.description as go_description
FROM smart_hr_backend_buobjectivegolink bgo
LEFT JOIN smart_hr_backend_buobjective bo ON bgo.objective_id = bo.id
LEFT JOIN smart_hr_backend_orgunit ou ON bo.org_unit_id = ou.id
LEFT JOIN smart_hr_backend_groupobjective go ON bgo.go_code_normalized = go.code
ORDER BY ou.name, bo.id
"""

# Execute queries
df_bu_objectives = pd.read_sql_query(bu_objectives_query, conn)
df_thrust_areas = pd.read_sql_query(thrust_areas_query, conn)
df_group_objectives = pd.read_sql_query(group_objectives_query, conn)
df_bu_ta_links = pd.read_sql_query(bu_ta_links_query, conn)
df_bu_go_links = pd.read_sql_query(bu_go_links_query, conn)

conn.close()

# Create Excel file with multiple sheets
output_file = 'DB_Dump_Verification.xlsx'

with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    df_bu_objectives.to_excel(writer, sheet_name='BU Objectives', index=False)
    df_thrust_areas.to_excel(writer, sheet_name='Thrust Areas', index=False)
    df_group_objectives.to_excel(writer, sheet_name='Group Objectives', index=False)
    df_bu_ta_links.to_excel(writer, sheet_name='BU-TA Links', index=False)
    df_bu_go_links.to_excel(writer, sheet_name='BU-GO Links', index=False)
    
    # Format the sheets
    for sheet_name in writer.sheets:
        ws = writer.sheets[sheet_name]
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Format header row
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

print(f"Database dump created successfully: {output_file}")
print(f"\nSummary:")
print(f"- Total BU Objectives: {len(df_bu_objectives)}")
print(f"- Total Thrust Areas: {len(df_thrust_areas)}")
print(f"- Total Group Objectives: {len(df_group_objectives)}")
print(f"- Total BU-TA Links: {len(df_bu_ta_links)}")
print(f"- Total BU-GO Links: {len(df_bu_go_links)}")
print(f"\nBU Objectives by Business Unit:")
print(df_bu_objectives['bu_name'].value_counts().sort_index())
